import argparse
import collections
import io
import json
import pathlib
import re
import sys
import time
import unicodedata
import urllib.parse
import urllib.request

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from config import load_key  # noqa: E402

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA, CACHE = ROOT / "data", ROOT / "data" / "cache" / "estat"
BASE = "https://api.e-stat.go.jp/rest/3.0/app/json"
STATS_CODE = "00450091"
UA = {"User-Agent": "AIIT-SP-Kadai4/1.0 (student coursework)"}

AGE_CODES = ["01", "02", "03", "04", "05", "06", "07",
             "08", "09", "10", "11", "12", "13"]

MEAN_COLS = {
    4: "年齢", 5: "勤続年数", 6: "所定内実労働時間数", 7: "超過実労働時間数",
    8: "きまって支給する現金給与額", 9: "所定内給与額",
    10: "年間賞与その他特別給与額", 11: "労働者数",
}
QUANTILE_ROWS = {-2: "1280", -1: "1290", 0: "1300", 1: "1310", 2: "1320"}


def norm(s):
    return re.sub(r"[\s　（）()0-9A-Za-z]", "",
                  unicodedata.normalize("NFKC", str(s)))


def industry_codes():
    rows = json.loads((DATA / "estat_ind.json").read_text(encoding="utf-8"))
    return {norm(r["name"]): r["code"] for r in rows}


def catalog(app_id):
    out, start = [], 1
    while start:
        q = urllib.parse.urlencode({"appId": app_id, "statsCode": STATS_CODE,
                                    "limit": 100, "startPosition": start})
        body = json.loads(urllib.request.urlopen(
            urllib.request.Request(f"{BASE}/getDataCatalog?{q}", headers=UA),
            timeout=300).read())["GET_DATA_CATALOG"]["DATA_CATALOG_LIST_INF"]
        inf = body["DATA_CATALOG_INF"]
        out += inf if isinstance(inf, list) else [inf]
        nxt = body["RESULT_INF"].get("NEXT_KEY")
        start = int(nxt) if nxt else None
        time.sleep(0.5)
    return out


def target_files(items, year):
    want = collections.defaultdict(list)
    for it in items:
        t = it["DATASET"]["TITLE"]
        if t.get("SURVEY_DATE") != year:
            continue
        if t.get("TABULATION_SUB_CATEGORY2") != "一般労働者":
            continue
        if t.get("TABULATION_SUB_CATEGORY3") not in ("産業大分類", "産業中分類"):
            continue
        res = it["RESOURCES"]["RESOURCE"]
        for r in (res if isinstance(res, list) else [res]):
            name = (r["TITLE"]["NAME"] if isinstance(r["TITLE"], dict)
                    else str(r["TITLE"]))
            if r.get("FORMAT") != "XLS_REP":
                continue
            if name.startswith("1_"):
                want["mean"].append((name, r["URL"]))
            elif name.startswith("3_"):
                want["quantile"].append((name, r["URL"]))
    return want


def download(url, name):
    key = re.search(r"statInfId=(\d+)", url)
    path = CACHE / f"{key.group(1) if key else abs(hash(url))}.xlsx"
    if path.exists():
        return path.read_bytes(), True
    raw = urllib.request.urlopen(
        urllib.request.Request(url, headers=UA), timeout=300).read()
    if raw[:2] != b"PK":
        raise RuntimeError(f"Excel ではない応答: {name}")
    path.write_bytes(raw)
    return raw, False


def sheet_industry(sheet_name, codes):
    return codes.get(norm(sheet_name.split("(")[0]))


def find_block(ws, label):
    for i in range(1, min(ws.max_row, 60) + 1):
        for c in range(2, 7):
            v = ws.cell(i, c).value
            if isinstance(v, str) and norm(v).startswith(label):
                return i
    return None


def to_number(v):
    blank = ("-", "－", "…", "***", "X", "")
    if v is None or (isinstance(v, str) and v.strip() in blank):
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def parse_mean(raw, codes):
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=False,
                                data_only=True)
    out = {}
    for name in wb.sheetnames:
        code = sheet_industry(name, codes)
        if code is None:
            continue
        ws = wb[name]
        head = find_block(ws, "男女計学歴計")
        if head is None:
            continue
        rows = {}
        for offset, age in enumerate(AGE_CODES):
            r = head + offset
            values = {label: to_number(ws.cell(r, col).value)
                      for col, label in MEAN_COLS.items()}
            if any(v is not None for v in values.values()):
                rows[age] = values
        if rows:
            out[code] = rows
    wb.close()
    return out


def parse_quantile(raw, codes):
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=False,
                                data_only=True)
    out = {}
    for name in wb.sheetnames:
        if "規模計" not in name:
            continue
        code = sheet_industry(name, codes)
        if code is None:
            continue
        ws = wb[name]
        head = find_block(ws, "男女計学歴計")
        if head is None:
            continue
        median = next((i for i in range(head, min(head + 60, ws.max_row + 1))
                       if any(isinstance(ws.cell(i, c).value, str)
                              and norm(ws.cell(i, c).value).startswith("中位数")
                              for c in range(2, 7))), None)
        if median is None:
            continue
        rows = collections.defaultdict(dict)
        for offset, qcode in QUANTILE_ROWS.items():
            for col, age in enumerate(AGE_CODES):
                v = to_number(ws.cell(median + offset, 6 + col).value)
                if v is not None:
                    rows[age][qcode] = v
        if rows:
            out[code] = dict(rows)
    wb.close()
    return out


def merge(target, addition):
    for code, rows in addition.items():
        target.setdefault(code, rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--years", type=int, nargs="+", default=[2024, 2025],
                    help="API に無い調査年（既定: 2024 2025）")
    args = ap.parse_args()

    CACHE.mkdir(parents=True, exist_ok=True)
    codes = industry_codes()
    print(f"産業分類 {len(codes)} 件を対応表に読み込み")
    print("統計表の一覧を取得中…")
    items = catalog(load_key("ESTAT_APP_ID"))
    print(f"  カタログ {len(items)} 件\n")

    mean_all = json.loads(
        (DATA / "estat_mean.json").read_text(encoding="utf-8"))
    quant_all = json.loads(
        (DATA / "estat_quantile.json").read_text(encoding="utf-8"))

    for year in args.years:
        files = target_files(items, year)
        if not files:
            print(f"{year}年: 統計表が見つからない")
            continue
        mean, quant, cached, started = {}, {}, 0, time.perf_counter()
        for kind, parse, sink in (("mean", parse_mean, mean),
                                  ("quantile", parse_quantile, quant)):
            for name, url in files[kind]:
                raw, hit = download(url, name)
                cached += hit
                merge(sink, parse(raw, codes))
        elapsed = time.perf_counter() - started
        n = sum(len(v) for v in files.values())
        print(f"{year}年  ファイル {n} 件（キャッシュ {cached}）  {elapsed:.1f}秒")
        print(f"        平均値 {len(mean)} 産業 / 分位数 {len(quant)} 産業")
        missing = set(mean) - set(quant)
        if missing:
            print(f"        分位数の無い産業 {len(missing)} 件")
        mean_all[str(year)] = mean
        quant_all[str(year)] = quant

    for path, obj in ((DATA / "estat_mean.json", mean_all),
                      (DATA / "estat_quantile.json", quant_all)):
        tmp = path.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(obj, ensure_ascii=False), encoding="utf-8")
        tmp.replace(path)
    print(f"\n収録年: 平均値 {sorted(mean_all)} / 分位数 {sorted(quant_all)}")


if __name__ == "__main__":
    main()
